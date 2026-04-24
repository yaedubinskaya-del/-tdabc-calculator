"""
web_app.py — веб-интерфейс TDABC-калькулятора «Стоимость рабочего времени»

Запуск:
    python run.py --web
    python run.py --web --share   # создаст публичную ссылку

Переменные окружения (необязательно, для Google Sheets):
    GOOGLE_CREDENTIALS_JSON  — содержимое JSON-файла сервисного аккаунта
    GOOGLE_SHEET_URL         — ссылка на Google Таблицу (должна быть открыта сервисному аккаунту)
"""

from __future__ import annotations

import io
import json
import os
import tempfile
from datetime import datetime

from calculator import calculate
from config import AssistantConfig, InsuranceRates, KpiPerProcedure, KpiRevenue, Procedure

# ═══════════════════════════════════════════════════════════
# ИНСТРУКЦИЯ
# ═══════════════════════════════════════════════════════════

INSTRUCTIONS = """
### Раздел «Сотрудник»

| Поле | Что вводить |
|------|------------|
| **ФИО сотрудника** | Фамилия и инициалы, например: Иванова А.А. |
| **Должность** | Операционная медсестра, ассистент врача, санитар и т.д. |

---

### Раздел «Оклад и страховые взносы»

| Поле | Что вводить |
|------|------------|
| **Оклад (gross), руб./мес.** | Оклад **до вычета НДФЛ** — именно эта сумма указана в трудовом договоре. Это не то, что сотрудник получает на руки |
| **Ставка страховых взносов, %** | Процент, который работодатель платит **сверх оклада** в ПФР, ФСС, ФОМС. Вводите одну цифру. Примеры: для МСП с суммы до МРОТ (88 000 руб.) — **30,2%**; свыше МРОТ — **15,1%**; если хотите усреднить — введите любую нужную цифру, например **22** |

> **Важно:** взносы начисляются также и на KPI. Калькулятор это учитывает автоматически.

---

### Раздел «График работы»

| Поле | Что вводить |
|------|------------|
| **Дней в неделю** | Сколько рабочих дней в неделю: 5 (пятидневка), 3 (сменный график) и т.д. Можно дробное: 2.5 |
| **Часов в день** | Длительность одной смены в часах: 8, 7, 12 и т.д. |
| **Непациентское время, мин/день** | Время в смене, когда сотрудник **не занят непосредственно пациентом**: обед, подготовка инструментов, уборка, ведение документации. Обычно **45–90 мин** |

> **Практическая мощность** — это ключевой знаменатель в формуле CCR:
> **Мощность = Дней/мес × (Часы × 60 − Непациентские мин)**
> Чем точнее вы укажете непациентское время, тем точнее будет стоимость минуты.

---

### Раздел «KPI» (необязательно)

KPI — переменная часть оплаты труда. Включите один или оба варианта, если они используются. Если KPI нет — оставьте флажки выключенными.

**KPI от выручки:**
| Поле | Что вводить |
|------|------------|
| **% от выручки** | Доля выручки подразделения, которую получает сотрудник как KPI. Пример: 0.006 означает 0,006% — при выручке 5 000 000 руб. KPI = 300 руб. |
| **Выручка подразделения, руб./мес.** | Плановая или фактическая выручка за месяц |

**KPI за процедуру:**
| Поле | Что вводить |
|------|------------|
| **Сумма за 1 процедуру, руб.** | Фиксированная сумма, которую сотрудник получает за каждую выполненную процедуру |
| **Кол-во процедур в месяц** | Плановое количество процедур |

---

### Раздел «Процедуры»

Каждая строка — одна процедура в формате:
```
Название процедуры;минуты;объём в месяц
```

**Пример:**
```
Гистероскопия;20;50
Большая операция;60;10
Перевязка;15;30
УЗИ органов малого таза;15;80
```

| Столбец | Что вводить |
|---------|------------|
| **Название** | Любое удобное название процедуры |
| **Минуты** | Время **непосредственного участия** этого сотрудника в процедуре (не общее время процедуры) |
| **Объём в месяц** | Плановое количество таких процедур за месяц — нужно для расчёта загрузки |

---

### Как читать результат

| Показатель | Смысл |
|-----------|-------|
| **CCR (руб./мин)** | **Стоимость 1 минуты работы сотрудника** — главный показатель для техкарты. Именно эту цифру вносят в таблицу расчёта стоимости услуг |
| **Стоимость процедуры, руб.** | CCR × минуты = сколько стоит участие этого сотрудника в одной процедуре |
| **Итого/мес, руб.** | Стоимость процедуры × объём в месяц |
| **Использование мощности, %** | Доля доступного времени, занятая плановыми процедурами. Норма: **70–85%**. Ниже 50% — сотрудник недозагружен, выше 90% — перегрузка |
"""


# ═══════════════════════════════════════════════════════════
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ═══════════════════════════════════════════════════════════

def _parse_procedures(text: str) -> list[Procedure]:
    """Парсит список процедур из многострочного ввода."""
    procedures: list[Procedure] = []
    for idx, raw_line in enumerate((text or "").splitlines(), start=1):
        line = raw_line.strip()
        if not line:
            continue
        parts = [p.strip() for p in line.split(";")]
        if len(parts) != 3:
            raise ValueError(
                f"Строка {idx}: используйте формат 'Название;минуты;объём'"
            )
        name = parts[0]
        try:
            minutes = float(parts[1].replace(",", "."))
            monthly_volume = int(float(parts[2].replace(",", ".")))
        except ValueError as exc:
            raise ValueError(
                f"Строка {idx}: минуты и объём должны быть числами"
            ) from exc
        if minutes < 0 or monthly_volume < 0:
            raise ValueError(f"Строка {idx}: минуты и объём не могут быть отрицательными")
        procedures.append(Procedure(name=name, minutes=minutes, monthly_volume=monthly_volume))
    if not procedures:
        raise ValueError("Добавьте хотя бы одну процедуру")
    return procedures


def _save_to_sheets(result, timestamp: str) -> str:
    """Сохранить строку результата в Google Таблицу. Возвращает статус."""
    creds_json = os.environ.get("GOOGLE_CREDENTIALS_JSON")
    sheet_url = os.environ.get("GOOGLE_SHEET_URL")
    if not creds_json or not sheet_url:
        return "ℹ️ Google Sheets не настроен (переменные окружения не заданы)"
    try:
        import gspread
        from google.oauth2.service_account import Credentials

        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        creds = Credentials.from_service_account_info(json.loads(creds_json), scopes=scopes)
        gc = gspread.authorize(creds)
        sh = gc.open_by_url(sheet_url)
        ws = sh.sheet1

        # Добавить заголовки, если лист пустой
        if not ws.get_all_values():
            ws.append_row([
                "Дата расчёта", "Сотрудник", "Должность",
                "Оклад, руб.", "Взносы на оклад, руб.", "KPI, руб.", "Взносы на KPI, руб.",
                "Итого ФОТ/мес, руб.", "Мощность, мин/мес", "CCR, руб./мин", "Загрузка, %",
            ])

        b = result.cost_breakdown
        c = result.capacity
        ws.append_row([
            timestamp, result.name, result.role,
            b.salary_gross, round(b.insurance_on_salary, 2), round(b.kpi_gross, 2), round(b.insurance_on_kpi, 2),
            round(b.total_monthly, 2), c.total_patient_min, round(result.ccr, 4),
            c.utilization_pct if c.utilization_pct is not None else "",
        ])
        return "✅ Данные сохранены в Google Таблицу"
    except Exception as e:
        return f"⚠️ Ошибка при записи в Google Sheets: {e}"


def _build_excel(result) -> bytes:
    """Создать Excel-файл с полным отчётом TDABC."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчёт TDABC"

    # Стили
    font_title   = Font(name="Times New Roman", bold=True, size=14)
    font_header  = Font(name="Times New Roman", bold=True, size=12)
    font_bold    = Font(name="Times New Roman", bold=True, size=11)
    font_normal  = Font(name="Times New Roman", size=11)
    fill_section = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_ccr     = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_thead   = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin         = Side(style="thin")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_right  = Alignment(horizontal="right")
    align_center = Alignment(horizontal="center")

    def section_row(r, text, fill=fill_section, cols=4):
        ws.merge_cells(f"A{r}:D{r}")
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = font_header
        cell.fill = fill
        cell.alignment = align_center

    def data_row(r, label, value, bold=False):
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = font_bold if bold else font_normal
        vc = ws.cell(row=r, column=2, value=value)
        vc.font = font_bold if bold else font_normal
        vc.alignment = align_right

    # ── Заголовок ──────────────────────────────────────────
    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1, value="Калькулятор стоимости рабочего времени (TDABC)")
    c.font = font_title
    c.alignment = align_center

    ws.cell(row=2, column=1,
            value=f"Расчёт от {datetime.now().strftime('%d.%m.%Y %H:%M')}").font = font_normal
    ws.merge_cells("A3:D3")
    c3 = ws.cell(row=3, column=1,
                 value=f"Сотрудник: {result.name}   |   Должность: {result.role}")
    c3.font = font_bold
    c3.alignment = align_center

    # ── Затраты ────────────────────────────────────────────
    row = 5
    section_row(row, "ЗАТРАТЫ")
    row += 1
    b = result.cost_breakdown
    items = [
        ("Оклад (gross), руб./мес.", b.salary_gross, False),
        ("Страховые взносы с оклада, руб.", b.insurance_on_salary, False),
        ("KPI (gross), руб.", b.kpi_gross, False),
        ("Взносы на KPI, руб.", b.insurance_on_kpi, False),
        ("ИТОГО для работодателя, руб./мес.", b.total_monthly, True),
    ]
    for label, val, bold in items:
        data_row(row, label, round(val, 2), bold=bold)
        row += 1

    # ── Мощность ───────────────────────────────────────────
    row += 1
    section_row(row, "ПРАКТИЧЕСКАЯ МОЩНОСТЬ")
    row += 1
    cap = result.capacity
    for label, val in [
        ("Рабочих дней/мес.", cap.days_per_month),
        ("Часов в смене", cap.hours_per_day),
        ("Непациентское время, мин/смену", cap.non_patient_min_per_day),
        ("Доступно пациентских мин/смену", cap.patient_min_per_day),
        ("Итого доступных мин/мес.", cap.total_patient_min),
    ]:
        data_row(row, label, val)
        row += 1

    # ── CCR ────────────────────────────────────────────────
    row += 1
    section_row(row, "CCR — СТОИМОСТЬ 1 МИНУТЫ РАБОТЫ", fill=fill_ccr)
    row += 1
    lc = ws.cell(row=row, column=1, value="CCR = Затраты ÷ Практическая мощность")
    lc.font = Font(name="Times New Roman", bold=True, size=12)
    vc = ws.cell(row=row, column=2, value=f"{result.ccr:.4f} руб./мин")
    vc.font = Font(name="Times New Roman", bold=True, size=12)
    vc.alignment = align_right
    row += 1
    if cap.utilization_pct is not None:
        flag = ""
        if cap.utilization_pct < 50:
            flag = "  ⚠ НЕДОЗАГРУЗКА"
        elif cap.utilization_pct > 90:
            flag = "  ⚠ ПЕРЕГРУЗКА"
        data_row(row, "Использование мощности, %", f"{cap.utilization_pct:.1f}%{flag}")
        row += 1

    # ── Таблица процедур ───────────────────────────────────
    row += 1
    section_row(row, "СТОИМОСТЬ ПРОЦЕДУР")
    row += 1
    proc_headers = ["Процедура", "Мин", "Стоимость, руб.", "Объём/мес", "Итого/мес, руб."]
    for col, h in enumerate(proc_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = font_bold
        cell.fill = fill_thead
        cell.border = border
        cell.alignment = align_center
    row += 1
    for p in result.procedures:
        for col, val in enumerate([
            p.name, p.minutes, round(p.cost_rub, 2), p.monthly_volume, round(p.monthly_total_rub, 2)
        ], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = font_normal
            cell.border = border
            if col > 1:
                cell.alignment = align_right
        row += 1

    # ── Ширина столбцов ────────────────────────────────────
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _calculate_from_form(
    name: str,
    role: str,
    salary: float,
    insurance_rate: float,
    days_per_week: float,
    hours_per_day: float,
    non_patient_min_per_day: float,
    use_kpi_revenue: bool,
    kpi_revenue_pct: float,
    revenue_per_month: float,
    use_kpi_proc: bool,
    kpi_proc_amount: float,
    kpi_proc_count: int,
    procedures_text: str,
):
    try:
        procedures = _parse_procedures(procedures_text)
    except ValueError as e:
        return f"❌ Ошибка в процедурах: {e}", None

    insurance = InsuranceRates(mode="standard", rate_standard=insurance_rate / 100)

    kpi_revenue = None
    if use_kpi_revenue:
        kpi_revenue = KpiRevenue(pct_of_revenue=kpi_revenue_pct / 100, revenue_per_month=revenue_per_month)

    kpi_proc = None
    if use_kpi_proc:
        kpi_proc = KpiPerProcedure(
            amount_per_procedure=kpi_proc_amount,
            procedures_per_month=int(kpi_proc_count),
        )

    staff = AssistantConfig(
        name=name.strip() or "Без имени",
        role=role.strip() or "Сотрудник",
        salary=salary,
        insurance=insurance,
        days_per_week=days_per_week,
        hours_per_day=hours_per_day,
        non_patient_min_per_day=non_patient_min_per_day,
        kpi_revenue=kpi_revenue,
        kpi_per_procedure=kpi_proc,
        procedures=procedures,
    )

    try:
        result = calculate(staff)
    except Exception as e:
        return f"❌ Ошибка расчёта: {e}", None

    timestamp = datetime.now().strftime("%d.%m.%Y %H:%M")
    sheets_status = _save_to_sheets(result, timestamp)

    b = result.cost_breakdown
    cap = result.capacity

    md_lines = [
        f"## Результат: {result.role} — {result.name}",
        f"*Расчёт от {timestamp}*",
        "",
        "### Затраты",
        "| Статья | Сумма, руб. |",
        "|--------|------------:|",
        f"| Оклад (gross) | {b.salary_gross:,.2f} |",
        f"| Страховые взносы с оклада | {b.insurance_on_salary:,.2f} |",
        f"| KPI (gross) | {b.kpi_gross:,.2f} |",
        f"| Взносы на KPI | {b.insurance_on_kpi:,.2f} |",
        f"| **ИТОГО для работодателя** | **{b.total_monthly:,.2f}** |",
        "",
        "### Практическая мощность",
        "| Показатель | Значение |",
        "|-----------|----------|",
        f"| Рабочих дней/мес. | {cap.days_per_month:.1f} |",
        f"| Доступных пациентских мин/смену | {cap.patient_min_per_day:.1f} |",
        f"| Итого доступных мин/мес. | {cap.total_patient_min:.0f} |",
        "",
        "---",
        "### 🎯 CCR — стоимость 1 минуты работы",
        f"# **{result.ccr:.4f} руб./мин**",
        "",
        f"*Режим KPI: {result.kpi_mode}*",
        "",
    ]

    if cap.utilization_pct is not None:
        if cap.utilization_pct < 50:
            flag = " ⚠️ НЕДОЗАГРУЗКА"
        elif cap.utilization_pct > 90:
            flag = " ⚠️ ПЕРЕГРУЗКА"
        else:
            flag = " ✅"
        md_lines.append(f"**Использование мощности: {cap.utilization_pct:.1f}%{flag}**")
        md_lines.append("")

    md_lines += [
        "### Стоимость процедур",
        "| Процедура | Мин | Стоимость, руб. | Объём/мес | Итого/мес, руб. |",
        "|-----------|----:|----------------:|----------:|----------------:|",
    ]
    for p in result.procedures:
        md_lines.append(
            f"| {p.name} | {p.minutes:.0f} | {p.cost_rub:,.2f} | {p.monthly_volume} | {p.monthly_total_rub:,.2f} |"
        )

    md_lines += ["", "---", sheets_status]

    # Excel-файл
    excel_bytes = _build_excel(result)
    fname = f"TDABC_{result.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.write(excel_bytes)
    tmp.flush()
    tmp.close()

    return "\n".join(md_lines), tmp.name


# ═══════════════════════════════════════════════════════════
# ЗАПУСК ВЕБ-ИНТЕРФЕЙСА
# ═══════════════════════════════════════════════════════════

def launch_web(share: bool = False) -> None:
    try:
        import gradio as gr
    except ImportError as exc:
        raise SystemExit("Не найден пакет gradio. Установите: pip install gradio") from exc

    with gr.Blocks(
        title="Калькулятор стоимости рабочего времени",
        theme=gr.themes.Soft(),
    ) as demo:

        gr.Markdown(
            "# Калькулятор стоимости рабочего времени\n"
            "**Метод TDABC (Time-Driven Activity-Based Costing)**"
        )

        with gr.Accordion("📖 Инструкция по заполнению (нажмите, чтобы раскрыть)", open=False):
            gr.Markdown(INSTRUCTIONS)

        gr.Markdown("---")

        # ── Сотрудник ──────────────────────────────────────
        gr.Markdown("### 👤 Сотрудник")
        with gr.Row():
            name = gr.Textbox(label="ФИО сотрудника", placeholder="Иванова А.А.")
            role = gr.Textbox(label="Должность", placeholder="Операционная медсестра")

        # ── Оклад и взносы ─────────────────────────────────
        gr.Markdown("### 💰 Оклад и страховые взносы")
        with gr.Row():
            salary = gr.Number(label="Оклад (gross), руб./мес.", value=115_000)
            insurance_rate = gr.Number(
                label="Ставка страховых взносов, %",
                value=30.2,
                info="Введите нужную ставку вручную, например: 30.2 или 15.1 или 22",
            )

        # ── График ─────────────────────────────────────────
        gr.Markdown("### 📅 График работы")
        with gr.Row():
            days_per_week = gr.Number(label="Дней в неделю", value=5)
            hours_per_day = gr.Number(label="Часов в день", value=8)
            non_patient_min_per_day = gr.Number(
                label="Непациентское время, мин/день",
                value=45,
                info="Обед + уборка + подготовка + документация",
            )

        # ── KPI ────────────────────────────────────────────
        gr.Markdown("### 🎯 KPI (если не используется — оставьте флажки выключенными)")

        with gr.Group():
            gr.Markdown("**KPI от выручки подразделения**")
            with gr.Row():
                use_kpi_revenue = gr.Checkbox(label="Включить KPI от выручки", value=False)
                kpi_revenue_pct = gr.Number(
                    label="% от выручки (например: 0.006 = 0,006%)",
                    value=0.006,
                )
                revenue_per_month = gr.Number(
                    label="Выручка подразделения, руб./мес.",
                    value=4_750_000,
                )

        with gr.Group():
            gr.Markdown("**KPI за каждую выполненную процедуру**")
            with gr.Row():
                use_kpi_proc = gr.Checkbox(label="Включить KPI за процедуру", value=False)
                kpi_proc_amount = gr.Number(label="Сумма за 1 процедуру, руб.", value=300)
                kpi_proc_count = gr.Number(
                    label="Кол-во процедур/мес.", value=100, precision=0
                )

        # ── Процедуры ──────────────────────────────────────
        gr.Markdown("### 🏥 Процедуры")
        gr.Markdown(
            "Каждая строка: `Название процедуры;минуты участия сотрудника;объём в месяц`"
        )
        procedures_text = gr.Textbox(
            label="Список процедур",
            placeholder=(
                "Гистероскопия;20;50\n"
                "Большая операция;60;10\n"
                "Пункция / малая манипуляция;20;40\n"
                "Перевязка;15;30"
            ),
            lines=8,
        )

        calculate_btn = gr.Button("🧮 Рассчитать", variant="primary", size="lg")
        gr.Markdown("---")

        result_md = gr.Markdown()
        download_file = gr.File(label="📥 Скачать отчёт Excel", visible=False)

        def run_calc(
            name, role, salary, insurance_rate,
            days_per_week, hours_per_day, non_patient_min,
            use_kpi_rev, kpi_rev_pct, rev_month,
            use_kpi_proc_flag, kpi_proc_amt, kpi_proc_cnt,
            procedures_text,
        ):
            md, file_path = _calculate_from_form(
                name, role, salary, insurance_rate,
                days_per_week, hours_per_day, non_patient_min,
                use_kpi_rev, kpi_rev_pct, rev_month,
                use_kpi_proc_flag, kpi_proc_amt, kpi_proc_cnt,
                procedures_text,
            )
            if file_path is None:
                return md, gr.update(visible=False, value=None)
            return md, gr.update(visible=True, value=file_path)

        calculate_btn.click(
            fn=run_calc,
            inputs=[
                name, role, salary, insurance_rate,
                days_per_week, hours_per_day, non_patient_min_per_day,
                use_kpi_revenue, kpi_revenue_pct, revenue_per_month,
                use_kpi_proc, kpi_proc_amount, kpi_proc_count,
                procedures_text,
            ],
            outputs=[result_md, download_file],
        )

    port = int(os.getenv("PORT", "7860"))
    demo.launch(server_name="0.0.0.0", server_port=port, share=share)